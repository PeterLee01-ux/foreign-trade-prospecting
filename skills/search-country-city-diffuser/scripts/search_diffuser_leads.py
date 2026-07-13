#!/usr/bin/env python3
"""Collect country/city diffuser prospects from GeoNames and Overture Maps."""

from __future__ import annotations

import argparse
import collections
import csv
import datetime as dt
import json
import math
import re
import tempfile
import unicodedata
import urllib.request
import zipfile
from pathlib import Path


CATEGORIES = {
    "hotel": ("Hospitality", "Hotel and commercial-space scenting"),
    "resort": ("Hospitality", "Resort and hotel scenting"),
    "health_spa": ("Wellness", "Wellness and spa scenting"),
    "spas": ("Spa", "Spa scenting"),
    "medical_spa": ("Spa", "Medical spa scenting"),
    "day_spa": ("Spa", "Day spa scenting"),
    "aromatherapy": ("Wellness", "Direct aromatherapy relevance"),
    "skin_care": ("Beauty", "Beauty retail and treatment rooms"),
    "beauty_and_spa": ("Beauty", "Beauty and spa scenting"),
    "beauty_salon": ("Beauty", "Beauty salon scenting"),
    "real_estate": ("Real Estate", "Showroom and office scenting"),
    "real_estate_investment": ("Real Estate", "Commercial real estate"),
    "real_estate_agent": ("Real Estate", "Property presentation spaces"),
    "real_estate_service": ("Real Estate", "Real estate services"),
    "property_management": ("Property Management", "Shared-space scenting"),
    "retail": ("Retail", "Retail scenting"),
    "boutique": ("Luxury Goods", "Boutique scenting"),
    "department_store": ("Retail", "Department-store scenting"),
    "shopping_center": ("Retail", "Shopping-center scenting"),
    "perfume_store": ("Retail", "Direct fragrance retail"),
    "gift_shop": ("Retail", "Gift and fragrance retail"),
    "home_decor": ("Lifestyle", "Home and lifestyle retail"),
    "furniture_store": ("Lifestyle", "Home retail"),
    "furniture_accessory_store": ("Lifestyle", "Home accessories"),
    "interior_design": ("Interior Design", "Interior design projects"),
    "architect": ("Architecture", "Architecture and commercial design"),
    "architectural_designer": ("Architecture", "Architecture and commercial design"),
    "wholesale_store": ("Wholesale", "Wholesale channel"),
    "wholesaler": ("Wholesale", "Wholesale channel"),
    "furniture_wholesalers": ("Wholesale", "Home-goods wholesale"),
    "b2b_furniture_and_housewares": ("Wholesale", "B2B home goods"),
}

GENERIC_LOCAL_PARTS = {
    "info", "sale", "sales", "contact", "hello", "support", "admin", "office",
    "service", "customerservice", "help", "enquiry", "inquiry", "mail", "marketing",
    "reservations", "booking", "orders", "noreply", "none",
}
EMAIL_RE = re.compile(r"^[A-Z0-9._%+\-]+@[A-Z0-9.\-]+\.[A-Z]{2,}$", re.I)
KEYWORD_PATTERNS = {
    "scent": r"\bscent(?:ed|ing|s)?\b",
    "aroma": r"\baroma(?:s|tic)?\b",
    "aromatherapy": r"\baromatherapy\b",
    "fragrance": r"\bfragrance(?:s)?\b",
    "diffuser": r"\bdiffuser(?:s)?\b",
    "essential oil": r"\bessential oils?\b",
    "perfume": r"\bperfume(?:s|ry)?\b",
}


def norm(value: object) -> str:
    text = unicodedata.normalize("NFKD", str(value or "").casefold())
    return "".join(ch for ch in text if ch.isalnum())


def email_is_generic(email: str) -> bool:
    local = email.split("@", 1)[0].casefold()
    compact = re.sub(r"[._+\-]", "", local)
    compact = re.sub(r"\d+$", "", compact)
    return compact in GENERIC_LOCAL_PARTS


def national_digits(phone: str, dial_code: str) -> str:
    digits = re.sub(r"\D", "", phone or "")
    dial = re.sub(r"\D", "", dial_code or "")
    if dial and digits.startswith(dial):
        digits = digits[len(dial):]
    return digits.lstrip("0")


def is_public_phone(phone: str, dial_code: str, prefixes: tuple[str, ...]) -> bool:
    digits = national_digits(phone, dial_code)
    return any(digits.startswith(re.sub(r"\D", "", prefix)) for prefix in prefixes)


def distance_km(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    radius = 6371.0
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dp, dl = math.radians(lat2 - lat1), math.radians(lon2 - lon1)
    a = math.sin(dp / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dl / 2) ** 2
    return 2 * radius * math.asin(math.sqrt(a))


def download_geonames(country: str, cache_dir: Path) -> Path:
    target = cache_dir / f"{country}.zip"
    if not target.exists():
        cache_dir.mkdir(parents=True, exist_ok=True)
        url = f"https://download.geonames.org/export/dump/{country}.zip"
        print(f"Downloading {url}")
        urllib.request.urlretrieve(url, target)
    return target


def load_cities(zip_path: Path, country: str, pop_min: int, pop_max: int, exceptions: set[str]):
    all_points, eligible, aliases = [], [], collections.defaultdict(list)
    with zipfile.ZipFile(zip_path) as archive, archive.open(f"{country}.txt") as source:
        for raw in source:
            parts = raw.decode("utf-8").rstrip("\n").split("\t")
            if len(parts) < 15:
                continue
            try:
                lat, lon, population = float(parts[4]), float(parts[5]), int(parts[14] or 0)
            except ValueError:
                continue
            all_points.append((lat, lon))
            if parts[6] != "P":
                continue
            names = {parts[1], parts[2], *(parts[3] or "").split(",")}
            is_exception = any(norm(name) in exceptions for name in names)
            if not (pop_min <= population <= pop_max or is_exception):
                continue
            city = {
                "name": parts[1], "ascii_name": parts[2], "province_code": parts[10],
                "population": population, "latitude": lat, "longitude": lon,
                "is_exception": is_exception,
            }
            index = len(eligible)
            eligible.append(city)
            for name in names:
                if norm(name):
                    aliases[norm(name)].append(index)
    if not eligible:
        raise RuntimeError("No eligible GeoNames cities; check country, population range, and exceptions.")
    return all_points, eligible, aliases


def derive_bbox(points: list[tuple[float, float]]) -> tuple[float, float, float, float]:
    if not points:
        raise RuntimeError("GeoNames dump contained no coordinates.")
    lats = [p[0] for p in points]
    lons = [p[1] for p in points]
    return min(lons) - 0.25, max(lons) + 0.25, min(lats) - 0.25, max(lats) + 0.25


def query_overture(country: str, release: str, bbox: tuple[float, float, float, float]):
    try:
        import duckdb
    except ImportError as exc:
        raise RuntimeError("Install scripts/requirements.txt before running.") from exc
    category_sql = ",".join("'" + item.replace("'", "''") + "'" for item in CATEGORIES)
    path = f"s3://overturemaps-us-west-2/release/{release}/theme=places/type=place/*"
    xmin, xmax, ymin, ymax = bbox
    sql = f"""
    SELECT id, names.primary AS company, categories.primary AS category,
           list_extract(addresses, 1).freeform AS street_address,
           list_extract(addresses, 1).locality AS locality,
           list_extract(addresses, 1).region AS region,
           list_extract(addresses, 1).postcode AS postcode,
           list_extract(addresses, 1).country AS country,
           list_extract(websites, 1) AS website,
           list_extract(emails, 1) AS email,
           list_extract(phones, 1) AS phone,
           bbox.xmin AS longitude, bbox.ymin AS latitude, confidence
    FROM read_parquet('{path}')
    WHERE bbox.xmin BETWEEN {xmin} AND {xmax}
      AND bbox.ymin BETWEEN {ymin} AND {ymax}
      AND list_extract(addresses, 1).country = '{country}'
      AND categories.primary IN ({category_sql})
      AND array_length(websites) > 0
      AND array_length(emails) > 0
      AND array_length(phones) > 0
      AND coalesce(operating_status, 'open') <> 'closed'
    """
    con = duckdb.connect()
    con.execute("INSTALL httpfs; LOAD httpfs; SET s3_region='us-west-2';")
    result = con.execute(sql)
    columns = [col[0] for col in result.description]
    rows = [dict(zip(columns, row)) for row in result.fetchall()]
    rows.sort(key=lambda row: float(row.get("confidence") or 0), reverse=True)
    return rows


def match_city(record, cities, aliases, nearest_km: float):
    locality_key = norm(record.get("locality"))
    candidates = aliases.get(locality_key, [])
    if candidates:
        return min(candidates, key=lambda i: distance_km(record["latitude"], record["longitude"], cities[i]["latitude"], cities[i]["longitude"]))
    if locality_key:
        return None
    nearest = min(range(len(cities)), key=lambda i: distance_km(record["latitude"], record["longitude"], cities[i]["latitude"], cities[i]["longitude"]))
    km = distance_km(record["latitude"], record["longitude"], cities[nearest]["latitude"], cities[nearest]["longitude"])
    return nearest if km <= nearest_km else None


def street_name(address: str) -> str:
    value = (address or "").split(",")[0].strip()
    value = re.sub(r"^\d+[A-Za-z\-]*\s+", "", value)
    return re.sub(r"\s+(Suite|Ste|Unit|Floor|Fl)\s+.*$", "", value, flags=re.I).strip()


def filter_records(raw_rows, cities, aliases, args):
    selected, seen_ids, seen_keys = [], set(), set()
    prefixes = tuple(args.public_prefix)
    for record in raw_rows:
        email = str(record.get("email") or "").strip().lower()
        if not EMAIL_RE.match(email) or email_is_generic(email):
            continue
        if is_public_phone(str(record.get("phone") or ""), args.dial_code, prefixes):
            continue
        city_index = match_city(record, cities, aliases, args.nearest_km)
        if city_index is None:
            continue
        city = cities[city_index]
        key = (norm(record.get("company")), norm(city["ascii_name"]), norm(city["province_code"]))
        street = street_name(str(record.get("street_address") or ""))
        if not all((key[0], record.get("website"), record.get("phone"), street)):
            continue
        if record["id"] in seen_ids or key in seen_keys:
            continue
        company = str(record.get("company") or "")
        direct = [label for label, pattern in KEYWORD_PATTERNS.items() if re.search(pattern, company, re.I)]
        industry, use_case = CATEGORIES[record["category"]]
        record = dict(record)
        record.update({
            "city": city["name"], "city_ascii": city["ascii_name"],
            "province_code": city["province_code"], "population": city["population"],
            "city_exception": city["is_exception"], "cluster_street": street,
            "email": email, "email_status": "Public source + syntax passed; not SMTP verified",
            "phone_status": "Public-service prefixes excluded",
            "keyword_match": ", ".join(direct),
            "current_company_industry": industry, "business_use_case": use_case,
            "relevance_score": 50 + (20 if direct else 0) + 20 + (10 if float(record.get("confidence") or 0) >= 0.8 else 0),
        })
        selected.append(record)
        seen_ids.add(record["id"])
        seen_keys.add(key)
    for index, record in enumerate(selected, 1):
        record["record_no"] = index
    return selected


def quality_report(rows, remote_count: int, args) -> dict:
    ids = [row["id"] for row in rows]
    keys = [(norm(row["company"]), norm(row["city_ascii"]), norm(row["province_code"])) for row in rows]
    missing = sum(not all((r.get("company"), r.get("street_address"), r.get("website"), r.get("email"), r.get("phone"))) for r in rows)
    outside = sum(not r["city_exception"] and not args.population_min <= int(r["population"]) <= args.population_max for r in rows)
    return {
        "country": args.country, "country_name": args.country_name,
        "overture_release": args.release, "collected_at_utc": dt.datetime.now(dt.timezone.utc).isoformat(),
        "source_candidates": remote_count, "qualified_records": len(rows),
        "unique_ids": len(set(ids)), "duplicate_ids": len(ids) - len(set(ids)),
        "unique_business_keys": len(set(keys)), "duplicate_business_keys": len(keys) - len(set(keys)),
        "generic_emails": sum(email_is_generic(r["email"]) for r in rows),
        "public_numbers": sum(is_public_phone(r["phone"], args.dial_code, tuple(args.public_prefix)) for r in rows),
        "missing_required_fields": missing, "non_exception_population_violations": outside,
        "population_range": [args.population_min, args.population_max], "city_exceptions": args.include_city,
        "cities": dict(collections.Counter(r["city_ascii"] for r in rows)),
        "industries": dict(collections.Counter(r["current_company_industry"] for r in rows)),
        "regions": dict(collections.Counter(r["province_code"] for r in rows)),
        "claim_scope": "All records matching the stated filters in the cited public-data release.",
    }


def export_files(rows: list[dict], report: dict, output_dir: Path, args):
    output_dir.mkdir(parents=True, exist_ok=True)
    (output_dir / "qualified_leads.json").write_text(json.dumps(rows, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    fields = [
        "record_no", "country", "company", "current_company_industry", "category", "business_use_case",
        "keyword_match", "relevance_score", "city", "city_ascii", "city_exception", "population",
        "street_address", "cluster_street", "region", "province_code", "postcode", "website", "email",
        "email_status", "phone", "phone_status", "confidence", "id", "longitude", "latitude",
    ]
    with (output_dir / "qualified_leads.csv").open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, extrasaction="ignore")
        writer.writeheader(); writer.writerows(rows)
    (output_dir / "quality_report.json").write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    export_xlsx(rows, report, output_dir / "country_city_diffuser_leads.xlsx", fields, args)


def export_xlsx(rows, report, path: Path, fields: list[str], args):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError("Install scripts/requirements.txt before running.") from exc
    wb = Workbook(); notes = wb.active; notes.title = "Project Notes"
    leads = wb.create_sheet("Qualified Leads"); clusters = wb.create_sheet("City Street Clusters")
    stats = wb.create_sheet("Industry Region Stats"); rules = wb.create_sheet("Rules")
    title_fill = PatternFill("solid", fgColor="17365D")
    white_bold = Font(color="FFFFFF", bold=True)
    notes.append([f"{args.country_name} diffuser prospecting", "Result"])
    for key, value in report.items():
        notes.append([key, json.dumps(value, ensure_ascii=False) if isinstance(value, (dict, list)) else value])
    leads.append(fields)
    for row in rows: leads.append([row.get(field, "") for field in fields])
    counter = collections.Counter((r["cluster_street"], r["city_ascii"], r["province_code"]) for r in rows)
    clusters.append(["Street", "City", "Region Code", "Companies"])
    for key, count in sorted(counter.items(), key=lambda item: (-item[1], item[0])): clusters.append([*key, count])
    stats.append(["Dimension", "Value", "Count"])
    for dimension in ("current_company_industry", "province_code", "city_ascii"):
        for value, count in collections.Counter(r[dimension] for r in rows).most_common(): stats.append([dimension, value, count])
    rules.append(["Rule", "Value"])
    rules_data = [
        ("Country", args.country), ("Population", f"{args.population_min}-{args.population_max}"),
        ("Exceptions", ", ".join(args.include_city) or "None"), ("Calling code", args.dial_code),
        ("Public prefixes excluded", ", ".join(args.public_prefix)),
        ("Email", "Public source + syntax passed; generic role accounts excluded; not SMTP verified"),
        ("Deduplication", "Source ID and normalized company + city + region"),
        ("Overture", f"https://docs.overturemaps.org/ release {args.release}"),
        ("GeoNames", f"https://download.geonames.org/export/dump/{args.country}.zip"),
    ]
    for item in rules_data: rules.append(item)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]: cell.fill = title_fill; cell.font = white_bold; cell.alignment = Alignment(wrap_text=True)
        for row in ws.iter_rows(min_row=2):
            for cell in row: cell.alignment = Alignment(vertical="top", wrap_text=True)
        for index, column in enumerate(ws.columns, 1):
            width = min(45, max(10, max(len(str(cell.value or "")) for cell in list(column)[:200]) + 2))
            ws.column_dimensions[get_column_letter(index)].width = width
    wb.save(path)


def self_test():
    assert norm("  Dübai LLC ") == "dubaillc"
    assert email_is_generic("info@example.com")
    assert email_is_generic("sales-2@example.com")
    assert not email_is_generic("peter.lee@example.com")
    assert national_digits("+971 600 123456", "971") == "600123456"
    assert is_public_phone("+971 600 123456", "971", ("600",))
    assert not is_public_phone("+971 55 1234567", "971", ("600", "800"))
    assert 0 < distance_km(25.2, 55.3, 25.3, 55.3) < 20
    print("Self-test passed")


def parse_args():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--country", help="ISO alpha-2 code")
    parser.add_argument("--country-name", default="")
    parser.add_argument("--release", default="2026-06-17.0")
    parser.add_argument("--population-min", type=int, default=10_000)
    parser.add_argument("--population-max", type=int, default=250_000)
    parser.add_argument("--include-city", action="append", default=[])
    parser.add_argument("--dial-code", default="")
    parser.add_argument("--public-prefix", action="append", default=[])
    parser.add_argument("--nearest-km", type=float, default=20.0)
    parser.add_argument("--bbox", nargs=4, type=float, metavar=("XMIN", "XMAX", "YMIN", "YMAX"))
    parser.add_argument("--cache-dir", type=Path, default=Path(tempfile.gettempdir()) / "diffuser-geonames")
    parser.add_argument("--output-dir", type=Path, default=Path("outputs") / "country-city-diffuser")
    parser.add_argument("--self-test", action="store_true")
    return parser.parse_args()


def main():
    args = parse_args()
    if args.self_test:
        self_test(); return
    if not args.country:
        raise SystemExit("--country is required")
    args.country = args.country.upper()
    if not re.fullmatch(r"[A-Z]{2}", args.country):
        raise SystemExit("--country must be a two-letter ISO 3166-1 alpha-2 code")
    if not re.fullmatch(r"\d{4}-\d{2}-\d{2}\.\d+", args.release):
        raise SystemExit("--release must look like 2026-06-17.0")
    if args.population_min < 0 or args.population_max < args.population_min:
        raise SystemExit("population range is invalid")
    args.country_name = args.country_name or args.country
    exceptions = {norm(city) for city in args.include_city}
    geo_zip = download_geonames(args.country, args.cache_dir)
    points, cities, aliases = load_cities(geo_zip, args.country, args.population_min, args.population_max, exceptions)
    bbox = tuple(args.bbox) if args.bbox else derive_bbox(points)
    raw_rows = query_overture(args.country, args.release, bbox)
    rows = filter_records(raw_rows, cities, aliases, args)
    report = quality_report(rows, len(raw_rows), args)
    export_files(rows, report, args.output_dir, args)
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
